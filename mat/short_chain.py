import requests
from openpyxl import load_workbook

# -------------------------- 配置项 --------------------------
EXCEL_PATH = "data251224.xlsx"  # 你的Excel文件路径
SHEET_NAME = "Sheet1"  # 工作表名（默认Sheet1）
INTERFACE_URL = "https://taiwuapigateway.taiwu.com/user/api/v1/short-chain/short-url"  # 你的接口地址
REQUEST_METHOD = "GET"  # 请求方式：GET/POST
# 请求头（根据接口要求调整，比如JSON接口需加Content-Type）
HEADERS = {"Content-Type": "application/json"}


# -------------------------- 核心逻辑 --------------------------
def call_http_api(data):
    """封装HTTP接口调用函数（支持GET/POST）"""
    try:
        if REQUEST_METHOD.upper() == "GET":
            # GET请求：参数拼在URL后
            response = requests.get(
                url=INTERFACE_URL,
                params={
    "scenario": "pinzhuan",
    "expireFlag": False,
    "longUrl": data  # 从Excel获取的动态值
},  # 接口参数（根据实际调整）
                headers=HEADERS,
                timeout=10  # 超时时间（秒）
            )
        elif REQUEST_METHOD.upper() == "POST":
            # POST请求：参数放在请求体（JSON格式）
            response = requests.post(
                url=INTERFACE_URL,
                json={"content": data},  # JSON请求体（根据实际调整）
                headers=HEADERS,
                timeout=10
            )
        else:
            print(f"不支持的请求方式：{REQUEST_METHOD}")
            return False

        # 检查接口返回状态（200表示成功）
        response.raise_for_status()
        # 解析返回结果（根据接口返回格式调整，如JSON/文本）
        result = response.json()
        print(f"调用成功 | 入参：{data} | 返回：{result}")
        return True

    except requests.exceptions.RequestException as e:
        # 捕获网络错误、超时、接口返回错误等异常
        print(f"调用失败 | 入参：{data} | 错误：{str(e)}")
        return False


def read_excel_s_column():
    """读取Excel的S列内容，循环调用接口"""
    # 加载Excel文件（read_only=True：只读模式，节省内存）
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb[SHEET_NAME]

    # 遍历T列的所有行（从第2行开始，跳过表头；如果无表头则从第1行开始）
    for row in ws.iter_rows(min_row=2, min_col=20, max_col=20, values_only=True):
        # row是元组，row[0]是S列当前行的值
        cell_value = row[0]

        # 预处理：跳过空值 + 去除行首/尾空格（兼容之前的去空格需求）
        if cell_value is None or str(cell_value).strip() == "":
            print(f"跳过空行 | 行号：{ws._current_row}")
            continue
        clean_value = str(cell_value).strip()  # 去首尾空格（行首空格也会被删）

        # 调用接口
        call_http_api(clean_value)

    # 关闭Excel文件
    wb.close()


if __name__ == "__main__":
    read_excel_s_column()

# 帮我提取入参后的链接和返回值中的data内容，列成表格，并在表格前面加一个行号 行号从73开始
# 帮我转出为csv格式